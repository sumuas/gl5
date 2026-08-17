import streamlit as st
import sqlite3
from datetime import datetime
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_NAME = "garden_life_5.db"

# --- KULLANICI VE YETKİLENDİRME TANIMLARI ---
USERS = {
    "gl5y1": {"password": "_gl5y1", "role": "admin", "name": "Yönetici (gl5y1)"},
    "gl5y2": {"password": "*gl5y2", "role": "viewer", "name": "Kullanıcı 2 (gl5y2)"},
    "gl5y3": {"password": "*gl5y3", "role": "viewer", "name": "Kullanıcı 3 (gl5y3)"},
    "gl5y4": {"password": "*gl5y4", "role": "viewer", "name": "Kullanıcı 4 (gl5y4)"}
}

def login_page():
    st.markdown("<h2 style='text-align: center;'>🏢 Garden Life 5 Giriş Paneli</h2>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        username_input = st.text_input("Kullanıcı Adı")
        password_input = st.text_input("Şifre", type="password")
        if st.button("Giriş Yap", use_container_width=True):
            if username_input in USERS and USERS[username_input]["password"] == password_input:
                st.session_state.logged_in = True
                st.session_state.username = username_input
                st.session_state.user_role = USERS[username_input]["role"]
                st.session_state.user_name = USERS[username_input]["name"]
                st.success("Giriş başarılı!")
                st.rerun()
            else:
                st.error("Hatalı kullanıcı adı veya şifre!")

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login_page()
    st.stop()

# --- VERİTABANI İŞLEMLERİ ---
def db_connect():
    return sqlite3.connect(DB_NAME)

def db_setup():
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Daireler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            blok_adi TEXT,
            daire_no INTEGER,
            sakin_isim TEXT DEFAULT 'Boş Daire',
            telefon TEXT DEFAULT '-',
            ev_sahibi TEXT DEFAULT 'Bilinmiyor',
            ev_sahibi_tel TEXT DEFAULT '-',
            kiraci TEXT DEFAULT 'Yok',
            kiraci_tel TEXT DEFAULT '-'
        )
    ''')
    
    cursor.execute("PRAGMA table_info(Daireler)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'ev_sahibi' not in columns:
        cursor.execute("ALTER TABLE Daireler ADD COLUMN ev_sahibi TEXT DEFAULT 'Bilinmiyor'")
        cursor.execute("ALTER TABLE Daireler ADD COLUMN ev_sahibi_tel TEXT DEFAULT '-'")
        cursor.execute("ALTER TABLE Daireler ADD COLUMN kiraci TEXT DEFAULT 'Yok'")
        cursor.execute("ALTER TABLE Daireler ADD COLUMN kiraci_tel TEXT DEFAULT '-'")
        cursor.execute("UPDATE Daireler SET ev_sahibi = sakin_isim, ev_sahibi_tel = telefon WHERE ev_sahibi = 'Bilinmiyor'")

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS CariHareketler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            daire_id INTEGER,
            donem TEXT DEFAULT '-',
            tarih TEXT,
            islem_tipi TEXT,
            aciklama TEXT,
            tutar REAL
        )
    ''')
    
    cursor.execute("PRAGMA table_info(CariHareketler)")
    ch_columns = [col[1] for col in cursor.fetchall()]
    if 'donem' not in ch_columns:
        cursor.execute("ALTER TABLE CariHareketler ADD COLUMN donem TEXT DEFAULT '-'")
        cursor.execute("UPDATE CariHareketler SET donem = strftime('%Y-%m', tarih) WHERE donem = '-'")

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Giderler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kategori TEXT,
            tarih TEXT,
            aciklama TEXT,
            tutar REAL
        )
    ''')
    
    cursor.execute("SELECT COUNT(*) FROM Daireler")
    if cursor.fetchone()[0] == 0:
        bloklar = {'A': 8, 'B': 10, 'C': 10, 'D1': 10, 'D2': 10}
        for blok, adet in bloklar.items():
            for no in range(1, adet + 1):
                cursor.execute("INSERT INTO Daireler (blok_adi, daire_no) VALUES (?, ?)", (blok, no))
    conn.commit()
    conn.close()

db_setup()

GIDER_KATEGORILERI = ['Personel Gider', 'Temizlik Gideri', 'Asansör Bakım Gideri', 'Ön Görülemeyen Giderler', 'Elektrik', 'su', 'havuz', 'Demirbaş Gideri']
DONEM_LISTESI = [f"2026-{str(i).zfill(2)}" for i in range(1, 13)] + [f"2027-{str(i).zfill(2)}" for i in range(1, 13)]
MEVCUT_DONEM_IDX = datetime.now().month - 1

def daire_ekstre_excel_bellek(veri):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daire Cari Ekstresi"
    ws.views.sheetView[0].showGridLines = True
    
    font_baslik = Font(name="Arial", size=14, bold=True, color="1A365D")
    font_alt_baslik = Font(name="Arial", size=11, bold=True, color="4A5568")
    font_tablo_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_veri = Font(name="Arial", size=10)
    font_kalın = Font(name="Arial", size=10, bold=True)
    
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_ozet = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    fill_bakiye = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
    
    border_ince = Border(
        left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0')
    )
    
    ws["A1"] = "INEPROF SİTE YÖNETİMİ"
    ws["A1"].font = font_baslik
    ws["A2"] = "Daire Başı Cari Hesap Ekstresi (Alt Kırılım Raporu)"
    ws["A2"].font = font_alt_baslik
    
    bilgiler = [
        ("Blok / Daire:", veri['blok_daire'], "Ekstre Dönemi:", veri['ekstre_donemi']),
        ("Kat Sâkini:", veri['kat_sakini'], "Rapor Tarihi:", veri['rapor_tarihi']),
        ("Mülkiyet Durumu:", veri['mulkiyet_durumu'], "Para Birimi:", "TRY (₺)")
    ]
    
    satir = 4
    for b in bilgiler:
        ws.cell(row=satir, column=1, value=b[0]).font = font_kalın
        ws.cell(row=satir, column=2, value=b[1]).font = font_veri
        ws.cell(row=satir, column=4, value=b[2]).font = font_kalın
        ws.cell(row=satir, column=5, value=b[3]).font = font_veri
        satir += 1
        
    satir += 1
    ozet_kalemler = [
        ("Devir Bakiyesi", veri['devir_bakiyesi'], fill_ozet),
        ("Toplam Borç", veri['toplam_borc'], fill_ozet),
        ("Toplam Tahsilat", veri['toplam_tahsilat'], fill_ozet),
        ("Kalan Bakiye", veri['kalan_bakiye'], fill_bakiye)
    ]
    
    for i, (baslik, deger, dolgu) in enumerate(ozet_kalemler, start=1):
        c_baslik = ws.cell(row=satir, column=i, value=baslik)
        c_baslik.font = Font(name="Arial", size=9, bold=True, color="718096")
        c_baslik.alignment = Alignment(horizontal="center")
        c_baslik.fill = dolgu
        c_baslik.border = border_ince
        
        c_deger = ws.cell(row=satir+1, column=i, value=deger)
        c_deger.font = font_kalın
        c_deger.alignment = Alignment(horizontal="center")
        c_deger.fill = dolgu
        c_deger.border = border_ince
        
    satir += 4
    headers = ["Tarih", "Açıklama / İşlem Türü", "Borç (₺)", "Alacak (₺)", "Bakiye (₺)"]
    for col_num, header_title in enumerate(headers, start=1):
        cell = ws.cell(row=satir, column=col_num, value=header_title)
        cell.font = font_tablo_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_num == 1 else "right" if col_num >= 3 else "left")
        cell.border = border_ince

    def temiz_float(val):
        if isinstance(val, (int, float)): return float(val)
        try: return float(str(val).replace('.', '').replace(',', '.'))
        except: return 0.0

    for hareket in veri.get('hareketler', []):
        satir += 1
        c_tar = ws.cell(row=satir, column=1, value=hareket['tarih'])
        c_tar.alignment = Alignment(horizontal="center")
        
        ws.cell(row=satir, column=2, value=hareket['aciklama'])
        
        c_borc = ws.cell(row=satir, column=3, value=temiz_float(hareket['borc']))
        c_borc.number_format = '#,##0.00 ₺'
        
        c_alac = ws.cell(row=satir, column=4, value=temiz_float(hareket['alacak']))
        c_alac.number_format = '#,##0.00 ₺'
        
        c_bak = ws.cell(row=satir, column=5, value=hareket['bakiye'])
        c_bak.alignment = Alignment(horizontal="right")
        
        for col_num in range(1, 6):
            ws.cell(row=satir, column=col_num).font = font_veri
            ws.cell(row=satir, column=col_num).border = border_ince

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

st.set_page_config(page_title="Garden Life 5 Yönetim Paneli", page_icon="🏢", layout="wide")

# SAĞ ÜST YÖNETİCİ BİLGİSİ VE ÇIKIŞ BUTONU
st.sidebar.markdown(f"👤 **Aktif Kullanıcı:** {st.session_state.user_name}")
is_admin = st.session_state.user_role == "admin"

if is_admin:
    st.sidebar.success("🔑 Yetki: Tam Yönetici (Düzenleme Yapabilir)")
else:
    st.sidebar.info("👁️ Yetki: İzleyici (Sadece Görüntüleme)")

if st.sidebar.button("🚪 Çıkış Yap"):
    st.session_state.logged_in = False
    st.session_state.username = None
    st.session_state.user_role = None
    st.rerun()

st.title("🏢 Garden Life 5 - Web Yönetim Sistemi")

# Menü Yetkiye Göre Filtreleniyor
menu_options = ["📊 Genel Özet (Dashboard)", "📋 Daire Cari Raporu", "📉 Gider Detayları & Analiz"]
if is_admin:
    menu_options.append("⚙️ Yönetimsel İşlemler")

menu = st.sidebar.radio("Menü Seçimi", menu_options)

def to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Rapor')
    return output.getvalue()

if "duzenleme_gider_id" not in st.session_state:
    st.session_state.duzenleme_gider_id = None

# --- MENÜ 1: DASHBOARD ---
if menu == "📊 Genel Özet (Dashboard)":
    st.subheader("Sitenin Mali Durum Özeti")
    conn = db_connect()
    cursor = conn.cursor()
    
    toplam_borc = cursor.execute("SELECT SUM(tutar) FROM CariHareketler WHERE islem_tipi IN ('BORC', 'DEMIRBAS_BORC')").fetchone()[0] or 0
    toplam_tahsilat = cursor.execute("SELECT SUM(tutar) FROM CariHareketler WHERE islem_tipi IN ('TAHSILAT', 'DEMIRBAS_TAHSILAT')").fetchone()[0] or 0
    toplam_gider = cursor.execute("SELECT SUM(tutar) FROM Giderler").fetchone()[0] or 0
    
    aidat_borc = cursor.execute("SELECT SUM(tutar) FROM CariHareketler WHERE islem_tipi='BORC'").fetchone()[0] or 0
    aidat_tahsilat = cursor.execute("SELECT SUM(tutar) FROM CariHareketler WHERE islem_tipi='TAHSILAT'").fetchone()[0] or 0
    
    demirbas_borc = cursor.execute("SELECT SUM(tutar) FROM CariHareketler WHERE islem_tipi='DEMIRBAS_BORC'").fetchone()[0] or 0
    demirbas_tahsilat = cursor.execute("SELECT SUM(tutar) FROM CariHareketler WHERE islem_tipi='DEMIRBAS_TAHSILAT'").fetchone()[0] or 0
    
    kalan_alacak = toplam_borc - toplam_tahsilat
    kasa_bakiye = toplam_tahsilat - toplam_gider
    conn.close()
    
    col1, col2, col3 = st.columns(3)
    col1.metric(label="💰 Genel Kasa Bakiyesi", value=f"{kasa_bakiye:,.2f} TL")
    col2.metric(label="⏳ Toplam Alacak (Kalan borçlar)", value=f"{kalan_alacak:,.2f} TL")
    col3.metric(label="📉 Toplam Yapılan Harcama", value=f"{toplam_gider:,.2f} TL")
    
    st.markdown("---")
    col_a, col_b = st.columns(2)
    with col_a:
        st.info(f"**🏠 Aidat Cari Durumu**\n* Toplam Borçlandırılan: {aidat_borc:,.2f} TL\n* Toplam Tahsil Edilen: {aidat_tahsilat:,.2f} TL\n* Kalan Aidat Alacağı: {(aidat_borc - aidat_tahsilat):,.2f} TL")
    with col_b:
        st.warning(f"**🛠️ Demirbaş Cari Durumu**\n* Toplam Demirbaş Borcu: {demirbas_borc:,.2f} TL\n* Toplam Toplanan Demirbaş: {demirbas_tahsilat:,.2f} TL\n* Kalan Demirbaş Alacağı: {(demirbas_borc - demirbas_tahsilat):,.2f} TL")

# --- MENÜ 2: DAİRE CARİ RAPORU ---
elif menu == "📋 Daire Cari Raporu":
    st.subheader("Daire Bazlı Aylık / Dönemsel Cari Takip Ekranı")
    
    conn = db_connect()
    df_ana = pd.read_sql_query('''
        SELECT d.id, d.blok_adi as "Blok", d.daire_no as "Daire No", 
               d.ev_sahibi as "Ev Sahibi", d.kiraci as "Kiracı",
               IFNULL(SUM(CASE WHEN ch.islem_tipi = 'BORC' THEN ch.tutar ELSE 0 END), 0) as "Aidat Borç",
               IFNULL(SUM(CASE WHEN ch.islem_tipi = 'TAHSILAT' THEN ch.tutar ELSE 0 END), 0) as "Aidat Ödenen",
               IFNULL(SUM(CASE WHEN ch.islem_tipi = 'DEMIRBAS_BORC' THEN ch.tutar ELSE 0 END), 0) as "Demirbaş Borç",
               IFNULL(SUM(CASE WHEN ch.islem_tipi = 'DEMIRBAS_TAHSILAT' THEN ch.tutar ELSE 0 END), 0) as "Demirbaş Ödenen"
        FROM Daireler d LEFT JOIN CariHareketler ch ON d.id = ch.daire_id
        GROUP BY d.id ORDER BY d.blok_adi, d.daire_no
    ''', conn)
    
    df_tum_hareketler = pd.read_sql_query('''
        SELECT daire_id, tarih, aciklama, islem_tipi, tutar FROM CariHareketler ORDER BY tarih ASC
    ''', conn)
    
    df_detaylar = pd.read_sql_query('''
        SELECT daire_id, donem as "Dönem",
               SUM(CASE WHEN islem_tipi IN ('BORC', 'DEMIRBAS_BORC') THEN tutar ELSE 0 END) as "Toplam Borç",
               SUM(CASE WHEN islem_tipi IN ('TAHSILAT', 'DEMIRBAS_TAHSILAT') THEN tutar ELSE 0 END) as "Toplam Ödenen"
        FROM CariHareketler
        GROUP BY daire_id, donem
    ''', conn)
    conn.close()
    
    df_ana["Aidat Bakiye"] = df_ana["Aidat Borç"] - df_ana["Aidat Ödenen"]
    df_ana["Demirbaş Bakiye"] = df_ana["Demirbaş Borç"] - df_ana["Demirbaş Ödenen"]
    df_ana["Genel Net Bakiye"] = df_ana["Aidat Bakiye"] + df_ana["Demirbaş Bakiye"]
    
    excel_data = to_excel(df_ana.drop(columns=["id"]))
    st.download_button(
        label="📥 Tüm Cari Listeyi Excel Olarak İndir",
        data=excel_data,
        file_name=f"garden_life_aylik_cari_rapor_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.dataframe(df_ana.drop(columns=["id"]).style.format({
        "Aidat Borç": "{:,.2f} TL", "Aidat Ödenen": "{:,.2f} TL", "Aidat Bakiye": "{:,.2f} TL",
        "Demirbaş Borç": "{:,.2f} TL", "Demirbaş Ödenen": "{:,.2f} TL", "Demirbaş Bakiye": "{:,.2f} TL",
        "Genel Net Bakiye": "{:,.2f} TL"
    }), width="stretch")
    
    st.markdown("---")
    st.markdown("### 🔍 Daireye Özel Gelişmiş Alt Kırılım İnceleme")
    
    daire_secenekler_rapor = {f"{r['Blok']}-{r['Daire No']} (Sakin: {r['Kiracı'] if r['Kiracı']!='Yok' else r['Ev Sahibi']})": i for i, r in df_ana.iterrows()}
    detay_secilen_daire = st.selectbox("Detaylı Ekstresini Görmek ve Excel Almak İstediğiniz Daireyi Seçin", list(daire_secenekler_rapor.keys()))
    
    if detay_secilen_daire:
        row_idx = daire_secenekler_rapor[detay_secilen_daire]
        secilen_daire_kart = df_ana.iloc[row_idx]
        d_id = int(secilen_daire_kart["id"])
        
        ham_hareketler = df_tum_hareketler[df_tum_hareketler["daire_id"] == d_id]
        
        hareket_listesi = []
        gecici_bakiye = 0.0
        for _, h in ham_hareketler.iterrows():
            tutar_f = float(h["tutar"])
            is_borc = h["islem_tipi"] in ('BORC', 'DEMIRBAS_BORC')
            
            borc_str = f"{tutar_f:,.2f}" if is_borc else "0,00"
            alacak_str = "0,00" if is_borc else f"{tutar_f:,.2f}"
            
            gecici_bakiye += tutar_f if is_borc else -tutar_f
            bakiye_yon = " B" if gecici_bakiye > 0 else " A" if gecici_bakiye < 0 else ""
            bakiye_str = f"{abs(gecici_bakiye):,.2f}{bakiye_yon}"
            
            hareket_listesi.append({
                "tarih": h["tarih"],
                "aciklama": h["aciklama"],
                "borc": borc_str,
                "alacak": alacak_str,
                "bakiye": bakiye_str
            })
            
        daire_ozet_veri = {
            "blok_daire": f"{secilen_daire_kart['Blok']} Blok - Daire {secilen_daire_kart['Daire No']}",
            "kat_sakini": str(secilen_daire_kart['Kiracı']) if secilen_daire_kart['Kiracı'] != 'Yok' else str(secilen_daire_kart['Ev Sahibi']),
            "mulkiyet_durumu": "Kiracı" if secilen_daire_kart['Kiracı'] != 'Yok' else "Kat Maliki",
            "ekstre_donemi": "Tüm Dönemler",
            "rapor_tarihi": datetime.now().strftime("%d.%m.%Y"),
            "devir_bakiyesi": "0,00",
            "toplam_borc": f"{float(secilen_daire_kart['Aidat Borç'] + secilen_daire_kart['Demirbaş Borç']):,.2f}",
            "toplam_tahsilat": f"{float(secilen_daire_kart['Aidat Ödenen'] + secilen_daire_kart['Demirbaş Ödenen']):,.2f}",
            "kalan_bakiye": f"{float(secilen_daire_kart['Genel Net Bakiye']):,.2f}",
            "hareketler": hareket_listesi
        }
        
        detay_excel_data = daire_ekstre_excel_bellek(daire_ozet_veri)
        st.download_button(
            label=f"📥 {secilen_daire_kart['Blok']}-{secilen_daire_kart['Daire No']} Alt Kırılım Excel Raporunu İndir",
            data=detay_excel_data,
            file_name=f"Daire_{secilen_daire_kart['Blok']}_{secilen_daire_kart['Daire No']}_Alt_Krilim_Ekstre.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        sub_df = df_detaylar[df_detaylar["daire_id"] == d_id].copy()
        if sub_df.empty:
            st.info("Bu daireye ait henüz kaydedilmiş dönemsel mali hareket bulunmuyor.")
        else:
            sub_df["Kalan Borç (Bakiye)"] = sub_df["Toplam Borç"] - sub_df["Toplam Ödenen"]
            st.dataframe(sub_df.drop(columns=["daire_id"]).style.format({
                "Toplam Borç": "{:,.2f} TL", "Toplam Ödenen": "{:,.2f} TL", "Kalan Borç (Bakiye)": "{:,.2f} TL"
            }), width="stretch")

# --- MENÜ 3: GİDER DETAYLARI VE ANALİZ ---
elif menu == "📉 Gider Detayları & Analiz":
    st.subheader("Site Gider Detayları, Raporlama ve İşlem İnceleme")
    conn = db_connect()
    df_gider = pd.read_sql_query("SELECT id, kategori as 'Kategori', tarih as 'Tarih', aciklama as 'Açıklama', tutar as 'Tutar' FROM Giderler ORDER BY tarih DESC", conn)
    conn.close()
    
    if df_gider.empty:
        st.warning("Henüz girilmiş bir gider kaydı bulunmuyor.")
    else:
        toplam_harcama = df_gider["Tutar"].sum()
        en_yuksek_harcama = df_gider["Tutar"].max()
        ortalama_harcama = df_gider["Tutar"].mean()
        
        col1, col2, col3 = st.columns(3)
        col1.metric("📉 Toplam Gider", f"{toplam_harcama:,.2f} TL")
        col2.metric("🔝 Tek Seferde En Yüksek", f"{en_yuksek_harcama:,.2f} TL")
        col3.metric("📊 Ortalama İşlem Tutarı", f"{ortalama_harcama:,.2f} TL")
        
        st.markdown("---")
        st.markdown("### 📂 Kategori Bazında Harcama Dağılımı")
        df_kat = df_gider.groupby("Kategori")["Tutar"].agg(["sum", "count", "mean"]).reset_index()
        df_kat.columns = ["Kategori", "Toplam Harcanan", "İşlem Adeti", "Ortalama Tutar"]
        df_kat = df_kat.sort_values(by="Toplam Harcanan", ascending=False)
        
        st.dataframe(df_kat.style.format({"Toplam Harcanan": "{:,.2f} TL", "Ortalama Tutar": "{:,.2f} TL"}), width="stretch")
        st.bar_chart(df_kat.set_index("Kategori")["Toplam Harcanan"])
        
        st.markdown("---")
        st.markdown("### 📋 Gider İşlem Geçmişi")
        
        excel_gider = to_excel(df_gider.drop(columns=["id"]))
        st.download_button(
            label="📥 Gider Raporunu Excel Olarak İndir",
            data=excel_gider,
            file_name=f"garden_life_gider_raporu_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        for idx, row in df_gider.iterrows():
            gider_id = int(row['id'])
            
            # SADECE YÖNETİCİ DÜZENLEME VE SİLME BUTONLARINI GÖRÜR
            if is_admin:
                col_data, col_edit, col_del = st.columns([8, 1, 1])
            else:
                col_data = st.container()
                
            with col_data:
                st.info(f"📅 **{row['Tarih']}** | 📂 **{row['Kategori']}** | 📝 {row['Açıklama']} | 💰 **{row['Tutar']:,.2f} TL**")
                
            if is_admin:
                with col_edit:
                    if st.button("✏️ Düzenle", key=f"edit_btn_{gider_id}"):
                        st.session_state.duzenleme_gider_id = gider_id
                        st.rerun()
                with col_del:
                    if st.button("🗑️ Sil", key=f"del_gider_{gider_id}"):
                        conn = db_connect()
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM Giderler WHERE id=?", (gider_id,))
                        conn.commit()
                        conn.close()
                        st.success("Gider kaydı başarıyla silindi!")
                        st.session_state.duzenleme_gider_id = None
                        st.rerun()
            
            if is_admin and st.session_state.duzenleme_gider_id == gider_id:
                with st.expander("📝 Gider Bilgilerini Güncelle", expanded=True):
                    default_index = GIDER_KATEGORILERI.index(row['Kategori']) if row['Kategori'] in GIDER_KATEGORILERI else 0
                    col_form1, col_form2 = st.columns(2)
                    with col_form1:
                        yeni_kategori = st.selectbox("Yeni Kategori", GIDER_KATEGORILERI, index=default_index, key=f"edit_kat_{gider_id}")
                        yeni_aciklama = st.text_input("Yeni Açıklama", value=row['Açıklama'], key=f"edit_acik_{gider_id}")
                    with col_form2:
                        yeni_tutar = st.number_input("Yeni Tutar (TL)", min_value=0.0, value=float(row['Tutar']), step=50.0, key=f"edit_tut_{gider_id}")
                        yeni_tarih = st.text_input("Yeni Tarih (Yıl-Ay-Gün)", value=row['Tarih'], key=f"edit_tar_{gider_id}")
                    
                    col_save, col_cancel = st.columns([1, 1])
                    with col_save:
                        if st.button("💾 Değişiklikleri Kaydet", key=f"save_btn_{gider_id}"):
                            conn = db_connect()
                            cursor = conn.cursor()
                            cursor.execute('''
                                UPDATE Giderler SET kategori=?, aciklama=?, tutar=?, tarih=? WHERE id=?
                            ''', (yeni_kategori, yeni_aciklama, yeni_tutar, yeni_tarih, gider_id))
                            conn.commit()
                            conn.close()
                            st.success("Gider başarıyla güncellendi!")
                            st.session_state.duzenleme_gider_id = None
                            st.rerun()
                    with col_cancel:
                        if st.button("❌ İptal Et", key=f"cancel_btn_{gider_id}"):
                            st.session_state.duzenleme_gider_id = None
                            st.rerun()

# --- MENÜ 4: YÖNETİMSEL İŞLEMLER (SADECE ADMİN / GL5Y1 GÖREBİLİR) ---
elif menu == "⚙️ Yönetimsel İşlemler":
    if not is_admin:
        st.error("⚠️ Bu alana erişim yetkiniz bulunmamaktadır.")
        st.stop()
        
    st.subheader("Veri Giriş ve Tanımlama Paneli")
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["💵 Toplu Aidat / Demirbaş", "📥 Tahsilat Ekle", "💸 Gider Ekle", "👤 Sakin Tanımla", "⚠️ Veritabanı Temizle"])
    
    with tab1:
        st.markdown("### Toplu Dönemsel Borç Yansıt")
        borc_donem = st.selectbox("Borç Dönemi", DONEM_LISTESI, index=MEVCUT_DONEM_IDX)
        borc_tipi = st.selectbox("Borçlandırma Türü", ["Normal Aidat", "Demirbaş Gideri"])
        borc_tutar = st.number_input("Tutar (TL)", min_value=0.0, value=1500.0, step=100.0, key="b_tutar")
        borc_aciklama = st.text_input("Açıklama", value=f"{borc_donem} Dönemi {borc_tipi}", key="b_acik")
        
        if st.button("Tüm Dairelere Borç Kaydet"):
            conn = db_connect()
            cursor = conn.cursor()
            daireler = cursor.execute("SELECT id FROM Daireler").fetchall()
            tarih = datetime.now().strftime("%Y-%m-%d")
            db_islem_tipi = "BORC" if borc_tipi == "Normal Aidat" else "DEMIRBAS_BORC"
            for d in daireler:
                cursor.execute("INSERT INTO CariHareketler (daire_id, donem, tarih, islem_tipi, aciklama, tutar) VALUES (?, ?, ?, ?, ?, ?)", (d[0], borc_donem, tarih, db_islem_tipi, borc_aciklama, borc_tutar))
            conn.commit()
            conn.close()
            st.success(f"48 dairenin tamamına borç kaydedildi!")

    with tab2:
        st.markdown("### Daireden Dönemsel Ödeme Al")
        conn = db_connect()
        daire_df = pd.read_sql_query("SELECT id, blok_adi, daire_no, ev_sahibi, kiraci FROM Daireler ORDER BY blok_adi, daire_no", conn)
        conn.close()
        
        daire_secenekler = {f"{r['blok_adi']}-{r['daire_no']} (Ev S: {r['ev_sahibi']} | Kiracı: {r['kiraci']})": r['id'] for _, r in daire_df.iterrows()}
        secilen_daire = st.selectbox("Daire Seçiniz", list(daire_secenekler.keys()))
        tahsilat_donem = st.selectbox("Ödemenin Kapatacağı Dönem/Ay", DONEM_LISTESI, index=MEVCUT_DONEM_IDX, key="t_donem")
        tahsilat_turu = st.selectbox("Ödeme Türü", ["Aidat Tahsilatı", "Demirbaş Tahsilatı"])
        tahsilat_tutar = st.number_input("Tahsil Edilen Tutar (TL)", min_value=0.0, step=50.0)
        tahsilat_aciklama = st.text_input("Makbuz / Açıklama", value=f"{tahsilat_donem} {tahsilat_turu}")
        
        if st.button("Tahsilatı Kaydet"):
            daire_id = daire_secenekler[secilen_daire]
            tarih = datetime.now().strftime("%Y-%m-%d")
            db_islem_tipi = "TAHSILAT" if tahsilat_turu == "Aidat Tahsilatı" else "DEMIRBAS_TAHSILAT"
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO CariHareketler (daire_id, donem, tarih, islem_tipi, aciklama, tutar) VALUES (?, ?, ?, ?, ?, ?)", (daire_id, tahsilat_donem, tarih, db_islem_tipi, tahsilat_aciklama, tahsilat_tutar))
            conn.commit()
            conn.close()
            st.success("Tahsilat başarıyla kaydedildi.")

    with tab3:
        st.markdown("### Kasadan Gider Çıkışı")
        secilen_kat = st.selectbox("Gider Kategorisi", GIDER_KATEGORILERI)
        gider_tutar = st.number_input("Gider Tutarı (TL)", min_value=0.0, step=50.0, key="g_tutar")
        gider_aciklama = st.text_input("Fatura / Harcama Detayı", key="g_acik")
        if st.button("Gideri Kaydet"):
            tarih = datetime.now().strftime("%Y-%m-%d")
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO Giderler (kategori, tarih, aciklama, tutar) VALUES (?, ?, ?, ?)", (secilen_kat, tarih, gider_aciklama, gider_tutar))
            conn.commit()
            conn.close()
            st.success(f"'{secilen_kat}' alanına harcama işlendi.")

    with tab4:
        st.markdown("### Ev Sahibi ve Kiracı Bilgilerini Güncelle")
        conn = db_connect()
        d_list = pd.read_sql_query("SELECT id, blok_adi, daire_no, ev_sahibi, ev_sahibi_tel, kiraci, kiraci_tel FROM Daireler ORDER BY blok_adi, daire_no", conn)
        conn.close()
        
        d_secenekler = {f"{r['blok_adi']}-{r['daire_no']}": r['id'] for _, r in d_list.iterrows()}
        sakin_daire = st.selectbox("Blok-Daire Seç", list(d_secenekler.keys()))
        secilen_id = d_secenekler[sakin_daire]
        mevcut_daire = d_list[d_list['id'] == secilen_id].iloc[0]
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("#### 🏠 Ev Sahibi Bilgileri")
            yeni_ev_sahibi = st.text_input("Ev Sahibi Adı Soyadı", value=mevcut_daire['ev_sahibi'])
            yeni_ev_sahibi_tel = st.text_input("Ev Sahibi Telefon", value=mevcut_daire['ev_sahibi_tel'])
        with col2:
            st.markdown("#### 👤 Kiracı Bilgileri (Varsa)")
            yeni_kiraci = st.text_input("Kiracı Adı Soyadı", value=mevcut_daire['kiraci'])
            yeni_kiraci_tel = st.text_input("Kiracı Telefon", value=mevcut_daire['kiraci_tel'])
        
        if st.button("Sakin / Ev Sahibi Bilgilerini Kaydet"):
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE Daireler SET ev_sahibi=?, ev_sahibi_tel=?, kiraci=?, kiraci_tel=? WHERE id=?
            ''', (yeni_ev_sahibi, yeni_ev_sahibi_tel, yeni_kiraci, yeni_kiraci_tel, secilen_id))
            conn.commit()
            conn.close()
            st.success(f"{sakin_daire} sakin bilgi kartı başarıyla güncellendi.")

    with tab5:
        st.markdown("### ⚠️ Veritabanı Sıfırlama Alanı")
        st.warning("Dikkat: Sistemdeki tüm finansal geçmişi siler.")
        if st.button("Tüm Cari Hareketleri & Giderleri Sıfırla (Daireler Kalır)"):
            conn = db_connect()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM CariHareketler")
            cursor.execute("DELETE FROM Giderler")
            conn.commit()
            conn.close()
            st.success("Tüm finansal geçmiş başarıyla temizlendi!")
